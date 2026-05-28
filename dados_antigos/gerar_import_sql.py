"""
Gera SQL para importar dados do Excel antigo (Bd_gastos_viagem.xlsx) no Supabase.

Uso:
  python gerar_import_sql.py <CARTEIRA_ID>

Exemplo:
  python gerar_import_sql.py 550e8400-e29b-41d4-a716-446655440000

O UUID da carteira você pega rodando no SQL Editor do Supabase:
  SELECT id, nome FROM carteiras;

O script gera o arquivo import_supabase.sql na mesma pasta.
"""

import sys
import openpyxl
from datetime import datetime, date

def sql_str(v):
    """Escapa uma string para SQL."""
    if v is None:
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

def sql_ts(v):
    """Converte datetime para timestamp SQL."""
    if v is None:
        return 'NULL'
    if isinstance(v, datetime):
        return f"'{v.strftime('%Y-%m-%d %H:%M:%S')}+00'"
    if isinstance(v, date):
        return f"'{v.strftime('%Y-%m-%d')}'"
    return sql_str(v)

def sql_date(v):
    """Converte para date SQL."""
    if v is None:
        return 'NULL'
    if isinstance(v, (datetime, date)):
        return f"'{v.strftime('%Y-%m-%d')}'"
    return sql_str(v)

def sql_num(v):
    if v is None:
        return '0'
    return str(v)

def sql_bool(v):
    if v is None:
        return 'FALSE'
    return 'TRUE' if v else 'FALSE'

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    carteira_id = sys.argv[1].strip()

    wb = openpyxl.load_workbook('Bd_gastos_viagem.xlsx')

    lines = []
    lines.append("-- Import gerado automaticamente por gerar_import_sql.py")
    lines.append(f"-- Carteira alvo: {carteira_id}")
    lines.append(f"-- Gerado em: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")
    lines.append("BEGIN;")
    lines.append("")

    # --- VIAGENS ---
    ws = wb['Viagens']
    viagem_ids = []
    lines.append("-- =====================")
    lines.append("-- VIAGENS")
    lines.append("-- =====================")
    for row in ws.iter_rows(min_row=2, values_only=True):
        vid, nome, inicio, fim, orc_deb, orc_cred, criada_em, arquivada = row
        if vid is None:
            continue
        viagem_ids.append(vid)
        lines.append(
            f"INSERT INTO viagens (id, nome, inicio, fim, orcamento_debito, orcamento_credito, arquivada, created_at, carteira_id)"
            f" VALUES ({sql_str(vid)}, {sql_str(nome)}, {sql_date(inicio)}, {sql_date(fim)},"
            f" {sql_num(orc_deb)}, {sql_num(orc_cred)}, {sql_bool(arquivada)}, {sql_ts(criada_em)},"
            f" '{carteira_id}')"
            f" ON CONFLICT (id) DO NOTHING;"
        )
    lines.append(f"-- {len(viagem_ids)} viagem(ns) inserida(s)")
    lines.append("")

    # --- GASTOS ---
    ws = wb['Gastos']
    gastos_count = 0
    lines.append("-- =====================")
    lines.append("-- GASTOS")
    lines.append("-- =====================")
    for row in ws.iter_rows(min_row=2, values_only=True):
        gid, data, local, item, valor, tipo, data_gasto, categoria, status_remocao, solicitado_por, viagem_id, scope, _ = row
        if gid is None:
            continue
        gastos_count += 1
        lines.append(
            f"INSERT INTO gastos (id, viagem_id, scope, data, data_gasto, local, item, valor, tipo, categoria, status_remocao, solicitado_por, created_at, carteira_id)"
            f" VALUES ({sql_str(gid)}, {sql_str(viagem_id)}, {sql_str(scope or 'viagem')},"
            f" {sql_ts(data)}, {sql_date(data_gasto)}, {sql_str(local)}, {sql_str(item)},"
            f" {sql_num(valor)}, {sql_str(tipo)}, {sql_str(categoria or 'outros')},"
            f" {sql_str(status_remocao or '')}, {sql_str(solicitado_por)}, {sql_ts(data)},"
            f" '{carteira_id}')"
            f" ON CONFLICT (id) DO NOTHING;"
        )
    lines.append(f"-- {gastos_count} gasto(s) inserido(s)")
    lines.append("")

    # --- AJUSTES (Saldos) ---
    ws = wb['Saldos']
    ajustes_count = 0
    lines.append("-- =====================")
    lines.append("-- AJUSTES DE SALDO")
    lines.append("-- =====================")
    for row in ws.iter_rows(min_row=2, values_only=True):
        aid, data, tipo, valor_antes, valor_depois, motivo, por = row
        if aid is None:
            continue
        ajustes_count += 1
        lines.append(
            f"INSERT INTO ajustes (id, data, tipo, valor_antes, valor_depois, motivo, por, created_at, carteira_id)"
            f" VALUES ({sql_str(aid)}, {sql_ts(data)}, {sql_str(tipo)},"
            f" {sql_num(valor_antes)}, {sql_num(valor_depois)}, {sql_str(motivo or '')}, {sql_str(por)},"
            f" {sql_ts(data)}, '{carteira_id}')"
            f" ON CONFLICT (id) DO NOTHING;"
        )
    lines.append(f"-- {ajustes_count} ajuste(s) inserido(s)")
    lines.append("")
    lines.append("COMMIT;")

    out_file = 'import_supabase.sql'
    with open(out_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    print(f"Arquivo gerado: dados_antigos/{out_file}")
    print(f"  Viagens:  {len(viagem_ids)}")
    print(f"  Gastos:   {gastos_count}")
    print(f"  Ajustes:  {ajustes_count}")
    print()
    print("Próximo passo: cole o conteúdo de import_supabase.sql no SQL Editor do Supabase e execute.")

if __name__ == '__main__':
    main()
